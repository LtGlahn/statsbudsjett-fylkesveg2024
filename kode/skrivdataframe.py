"""
Skrivr (geodata)frame med større overskrift og kolonnebredde

Funker bare med openpyxl 3.0.5

Dette har vært et sant helvete å finne ut av. openpyxl er grei nok, men 
oppførselen har endret seg kraftig gjennom levetiden, så de flese eksemplene
er enten utdart eller nyere...  
"""
from datetime import datetime 
from copy import deepcopy
import pdb

import pandas as pd
import numpy as np 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side

def lagworksheet( dataFrame, ws, overskrift, sheet_name='Ark 1', transponerPerVegkategori=False): 

    if transponerPerVegkategori:
        dataFrame = transponerPerVegkategori( dataFrame ) 


    for r in dataframe_to_rows( dataFrame, index=False, header=True):
        ws.append(r)
   


    # bd = Side(style='thick', color="000000")
    # overskrift.border = Border(left=bd, top=bd, right=bd, bottom=bd)   

    for ii in range( 1, len( dataFrame.columns)+1):
        col = get_column_letter(ii)
        cell = ws.cell( row=1, column=ii) 
        cell.style = overskrift 
        ws.column_dimensions[ col ].width = 30
    
    ws.title = sheet_name



def transponerFylkePerVegkategori( df ): 
    """
    Konverterer rader med index fylke vegkategori lengde => index fylke E R F P K S 

    Klarer ikke vrenge hode rundt indekser, transponering og gruppering i dag... 

    Innkommende dataframe har vanlig plain index (dvs 0,1,... ) og kolonne fylke, vegkategori og lengde
    """

    if 'fylke' in df.columns and 'vegkategori' in df.columns and 'lengde' in df.columns: 

        data = []

        for junk, row in df.iterrows(): 
            data.append( {'fylke' : row['fylke'], row['vegkategori'] : round( row['lengde'] ), }   )

        df = pd.DataFrame( data )
        df = df.groupby( ['fylke']).sum().reset_index()

    else: 
        print( 'Kan ikke transponere denne dataframen per vegkategori. Kolonner=', df.columns) 

    return df 


def transponerKommunePerVegkategori( df ): 
    """
    Konverterer rader med index fylke kommune vegkategori lengde => index fylke kommune E R F P K S 

    Klarer ikke vrenge hode rundt indekser, transponering og gruppering i dag... 

    Innkommende dataframe har vanlig plain index (dvs 0,1,... ) og kolonne fylke, kommune, vegkategori og lengde
    """

    if 'fylke' in df.columns and 'kommune' in df.columns and 'vegkategori' in df.columns and 'lengde' in df.columns: 
        
        data  = []
        for junk, row in df.iterrows(): 

            data.append( {'fylke' : row['fylke'], 'kommune' : row['kommune'],  row['vegkategori'] : round( row['lengde'] ), }   )
            
        df = pd.DataFrame( data )
        df = df.groupby( ['fylke', 'kommune']).sum().reset_index()

    else: 
        print( 'Kan ikke transponere denne dataframen per kommune og vegkategori. Kolonner=', df.columns) 

    return df 

def fylkesnr2fylkesnavn( df, fylkesnrkolonne='fylke'):
    """
    Bytter ut fylkesnummer med fylkesnavn i dataframe 

    ARGUMENTS: 
        dataframe  - (geo)dataframe

    KEYWORDS:
        fylkesnrkolonne - string, navn på kolonnen med fylkesnummeret. default='fylke' 

    RETURNS 
        dataframe (tar kopi av orginalen og modifiserer kopien )
    """
    fylkesnavn = {  11: 'Rogaland',
                    54: 'Troms og Finnmark',
                    18: 'Nordland',
                    15: 'Møre og Romsdal',
                    34: 'Innlandet',
                    42: 'Agder',
                    38: 'Vestfold og Telemark',
                    46: 'Vestland',
                    30: 'Viken',
                    3: 'Oslo',
                    50: 'Trøndelag'}

    df = deepcopy( df )
    df[fylkesnrkolonne] = df[fylkesnrkolonne].apply( lambda x : fylkesnavn[x])
    return df 
    



def skrivdf2xlsx( dataFrame, filnavn, sheet_name='Ark 1', metadata=None, transponerPerVegkategori=False ): 
    """
    Lagrer (geodataframe) til excel med fornuftig formattering; Funker bare med openpyxl 3.0.5

    Fjerner kolonnene 'geometry' og 'geometri' hvis de finnes

    ARGUMENTS
        dataFrame = En enkelt (geo)dataframe, eller liste med slike 
    """
    wb = Workbook()
    dfliste = []
    navneliste = []

    if isinstance( dataFrame, list): 
        dfliste = dataFrame 
        dataFrame = dfliste[0]

    if isinstance( sheet_name, list): 
        navneliste = sheet_name 
        sheet_name = navneliste[0]

    # Sikrer oss at vi jobber med kopi
    data = pd.DataFrame( dataFrame )
    data = data.copy()


    # 
    overskrift = NamedStyle(name="overskrift")
    overskrift.font = Font(bold=True, size=14)

    if 'geometry' in data.columns: 
        data.drop( 'geometry', axis='columns', inplace=True)

    if 'geometri' in data.columns: 
        data.drop( 'geometri', axis='columns', inplace=True)

    ws = wb.active

    # iterer over liste med (geo)dataframes...  
    lagworksheet( data, ws, overskrift, sheet_name=sheet_name, transponerPerVegkategori=transponerPerVegkategori)
    if len( dfliste ) > 1: 
        for idx, dataFrame in enumerate(dfliste[1:]): 
            ws = wb.create_sheet()
            if len( navneliste) > idx+1: 
                sheet_name = navneliste[idx+1] 
            else: 
                sheet_name = 'Ark ' + str( idx+1 )

            lagworksheet( dataFrame, ws, overskrift, sheet_name=sheet_name)

    if metadata: 
        # ws = wb.create_sheet( title='Metadata')
        ws = wb.create_sheet( )
        a = [ (' ', ' '), ('Dato lagret', datetime.now().strftime( '%Y-%m-%d') ) ]
        a.extend( [ (x, metadata[x]) for x in metadata  ])
        df =  pd.DataFrame( a, columns=['Metadata', ' '])
        lagworksheet( df, ws, overskrift, sheet_name='Metadata' )

    wb.save( filnavn )

